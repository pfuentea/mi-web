from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate
from django.db import transaction
from django.db.models import Sum, F, Count, Q
from django.db.models.functions import Coalesce
from django.utils import timezone
from .models import Student, Activity, FundDistribution, Cuota, PagoCuota, Curso, CursoMembresia, Abono, Objetivo, ObjetivoAlumno, Gasto
from django.contrib.auth.models import User
from django.contrib import messages
from .access import get_current_curso, get_user_cursos, can_manage_curso, gestor_required, is_tesorero_principal


def registro(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        curso_name = request.POST.get('curso_name', '').strip()
        curso_year = request.POST.get('curso_year', '').strip()
        curso_description = request.POST.get('curso_description', '').strip()

        errors = []
        if not username or not password:
            errors.append('Usuario y contraseña son requeridos.')
        elif User.objects.filter(username=username).exists():
            errors.append('Ese nombre de usuario ya está en uso.')
        if not curso_name:
            errors.append('El nombre del curso es requerido.')
        if not curso_year or not curso_year.isdigit():
            errors.append('El año del curso debe ser un número válido.')

        if not errors:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username, password=password,
                    first_name=first_name, last_name=last_name, email=email,
                )
                curso = Curso.objects.create(
                    name=curso_name, year=int(curso_year),
                    description=curso_description,
                )
                CursoMembresia.objects.create(
                    user=user, curso=curso,
                    rol=CursoMembresia.ROL_TESORERO, is_principal=True,
                )
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                request.session['current_curso_id'] = curso.id
                messages.success(request, f'Bienvenido, {user.first_name or user.username}! Tu curso "{curso}" fue creado.')
                return redirect('admin_home')

        for error in errors:
            messages.error(request, error)

    return render(request, 'fondos/registro.html', {})


@login_required
def dashboard(request):
    if request.user.is_staff:
        return redirect('select_curso')
    # Check if user is tesorero/ayudante of any curso
    gestioned_cursos = get_user_cursos(request.user)
    if gestioned_cursos.exists():
        if gestioned_cursos.count() == 1:
            request.session['current_curso_id'] = gestioned_cursos.first().id
            return redirect('admin_home')
        return redirect('select_curso')
    # Regular apoderado — show their students
    students = Student.objects.filter(parent=request.user).select_related('curso').prefetch_related('distributions__activity__objetivo')
    students_data = []
    for s in students:
        meta = s.curso.meta_por_alumno if s.curso else None
        falta = max(0, meta - s.total_funds) if meta else None
        porcentaje = min(100, int(s.total_funds * 100 / meta)) if meta and meta > 0 else None
        students_data.append({
            'student': s, 'meta': meta, 'falta': falta, 'porcentaje': porcentaje,
            'objetivos_alumno': _get_objetivos_alumno(s),
        })
    context = {'students_data': students_data}
    return render(request, 'fondos/dashboard.html', context)


@gestor_required
def edit_curso(request):
    current_curso = get_current_curso(request)
    if not (request.user.is_staff or is_tesorero_principal(request.user, current_curso)):
        messages.error(request, 'Solo el administrador o el Tesorero Principal pueden editar el curso.')
        return redirect('admin_home')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        year = request.POST.get('year', '').strip()
        description = request.POST.get('description', '').strip()
        tipo = request.POST.get('tipo', '').strip()
        meta_str = request.POST.get('meta_por_alumno', '').strip()
        if name and year and year.isdigit():
            current_curso.name = name
            current_curso.year = int(year)
            current_curso.description = description
            current_curso.meta_por_alumno = int(meta_str) if meta_str.isdigit() and int(meta_str) > 0 else None
            if request.user.is_staff and tipo in (Curso.TIPO_NORMAL, Curso.TIPO_AVANZADO):
                current_curso.tipo = tipo
            current_curso.save()
            messages.success(request, f'Curso actualizado: {current_curso}')
        else:
            messages.error(request, 'Nombre y año son requeridos.')
    return redirect('admin_home')


@login_required
def select_curso(request):
    cursos = get_user_cursos(request.user)

    if request.method == 'POST' and request.user.is_staff:
        name = request.POST.get('name', '').strip()
        year = request.POST.get('year', '').strip()
        description = request.POST.get('description', '')
        tipo = request.POST.get('tipo', Curso.TIPO_NORMAL)
        if tipo not in (Curso.TIPO_NORMAL, Curso.TIPO_AVANZADO):
            tipo = Curso.TIPO_NORMAL
        if name and year and year.isdigit():
            curso = Curso.objects.create(name=name, year=int(year), description=description, tipo=tipo)
            request.session['current_curso_id'] = curso.id
            messages.success(request, f'Curso "{curso}" creado exitosamente.')
            return redirect('admin_home')
        else:
            messages.error(request, 'Nombre y año son requeridos.')

    context = {'cursos': cursos}
    return render(request, 'fondos/select_curso.html', context)


@login_required
def switch_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id, is_active=True)
    if request.user.is_staff or can_manage_curso(request.user, curso):
        request.session['current_curso_id'] = curso_id
        messages.success(request, f'Ahora gestionando: {curso}')
    return redirect('admin_home')


@gestor_required
def admin_home(request):
    current_curso = get_current_curso(request)
    alumnos = Student.objects.filter(curso=current_curso).count()
    apoderados = CursoMembresia.objects.filter(curso=current_curso).count()
    actividades = Activity.objects.filter(curso=current_curso).count()
    cuotas = Cuota.objects.filter(curso=current_curso).count()
    cuotas_pagadas = PagoCuota.objects.filter(cuota__curso=current_curso, paid=True).count()
    cuotas_pendientes = PagoCuota.objects.filter(cuota__curso=current_curso, paid=False).count()

    total_dist = FundDistribution.objects.filter(
        activity__curso=current_curso
    ).aggregate(t=Sum('amount'))['t'] or 0
    total_cuotas_monto = PagoCuota.objects.filter(
        cuota__curso=current_curso, paid=True
    ).annotate(monto=F('cuota__amount')).aggregate(t=Sum('monto'))['t'] or 0
    monto_promedio = (total_dist + total_cuotas_monto) // alumnos if alumnos else 0

    es_principal = request.user.is_staff or is_tesorero_principal(request.user, current_curso)
    context = {
        'alumnos': alumnos,
        'apoderados': apoderados,
        'actividades': actividades,
        'cuotas': cuotas,
        'cuotas_pagadas': cuotas_pagadas,
        'cuotas_pendientes': cuotas_pendientes,
        'monto_promedio': monto_promedio,
        'es_principal': es_principal,
    }
    return render(request, 'fondos/admin_home.html', context)


@gestor_required
def admin_dashboard(request):
    current_curso = get_current_curso(request)
    activities = Activity.objects.filter(curso=current_curso).annotate(
        monto_repartido=Coalesce(Sum('distributions__amount'), 0),
        queda_por_repartir=F('total_amount') - Coalesce(Sum('distributions__amount'), 0),
    ).order_by('-date')

    if request.method == 'POST':
        name = request.POST.get('name')
        date = request.POST.get('date')
        description = request.POST.get('description', '')
        total_amount = request.POST.get('total_amount', 0)
        objetivo_id = request.POST.get('objetivo_id') or None
        if name and date:
            Activity.objects.create(
                name=name, date=date, description=description,
                total_amount=total_amount, curso=current_curso,
                objetivo_id=objetivo_id,
            )
            messages.success(request, 'Actividad creada con éxito.')
            return redirect('admin_dashboard')

    objetivos = Objetivo.objects.filter(curso=current_curso)
    context = {'activities': activities, 'objetivos': objetivos}
    return render(request, 'fondos/admin_dashboard.html', context)


@gestor_required
def distribute_funds(request, activity_id):
    current_curso = get_current_curso(request)
    activity = get_object_or_404(Activity, id=activity_id, curso=current_curso)
    sort_by = request.GET.get('sort', 'name')
    sort_order = request.GET.get('order', 'asc')

    if sort_by == 'parent':
        students = Student.objects.filter(curso=current_curso).order_by(
            'parent__last_name' if sort_order == 'asc' else '-parent__last_name'
        )
    else:
        students = Student.objects.filter(curso=current_curso).order_by(
            ('last_name' if sort_order == 'asc' else '-last_name'), 'first_name'
        )

    if request.method == 'POST':
        selected_students_ids = request.POST.getlist('students')
        distribution_type = request.POST.get('distribution_type')

        if not selected_students_ids:
            messages.error(request, 'Debe seleccionar al menos un alumno.')
            return redirect('distribute_funds', activity_id=activity.id)

        selected_students = Student.objects.filter(id__in=selected_students_ids, curso=current_curso)

        if distribution_type == 'equal':
            total_amount = float(request.POST.get('total_amount', 0))
            if total_amount <= 0:
                messages.error(request, 'El monto a repartir debe ser mayor a 0.')
                return redirect('distribute_funds', activity_id=activity.id)
            amount_per_student = int(total_amount / len(selected_students))
            for student in selected_students:
                FundDistribution.objects.create(
                    student=student, activity=activity, amount=amount_per_student
                )
            messages.success(request, f'Dinero repartido equitativamente: ${amount_per_student} a {len(selected_students)} alumnos.')

        elif distribution_type == 'manual':
            assigned_count = 0
            for student in selected_students:
                amount_str = request.POST.get(f'amount_{student.id}', '0')
                if amount_str and amount_str.isdigit() and int(amount_str) > 0:
                    FundDistribution.objects.create(
                        student=student, activity=activity, amount=int(amount_str)
                    )
                    assigned_count += 1
            messages.success(request, f'Dinero asignado manualmente a {assigned_count} alumnos.')

        return redirect('admin_dashboard')

    existing_distributions = FundDistribution.objects.filter(activity=activity).select_related('student__parent')
    distributed_ids = list(existing_distributions.values_list('student_id', flat=True))
    total_distribuido = sum(d.amount for d in existing_distributions)

    context = {
        'activity': activity,
        'students': students,
        'existing_distributions': existing_distributions,
        'distributed_ids': distributed_ids,
        'total_distribuido': total_distribuido,
        'sort_by': sort_by,
        'sort_order': sort_order,
    }
    return render(request, 'fondos/distribute_funds.html', context)


@gestor_required
def manage_users(request):
    current_curso = get_current_curso(request)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_user':
            username = request.POST.get('username')
            password = request.POST.get('password')
            first_name = request.POST.get('first_name', '')
            last_name = request.POST.get('last_name', '')
            email = request.POST.get('email', '')
            if username and password:
                if User.objects.filter(username=username).exists():
                    messages.error(request, 'Ese nombre de usuario ya existe.')
                else:
                    user = User.objects.create_user(
                        username=username, password=password,
                        first_name=first_name, last_name=last_name, email=email,
                    )
                    if current_curso:
                        CursoMembresia.objects.create(user=user, curso=current_curso, rol=CursoMembresia.ROL_APODERADO)
                    messages.success(request, f'Apoderado {username} creado exitosamente.')
            else:
                messages.error(request, 'Usuario y contraseña son requeridos.')

        elif action == 'create_student':
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            parent_id = request.POST.get('parent_id')
            if first_name and last_name:
                parent = User.objects.filter(id=parent_id).first() if parent_id else None
                Student.objects.create(
                    first_name=first_name, last_name=last_name,
                    parent=parent, curso=current_curso,
                )
                if parent and current_curso:
                    CursoMembresia.objects.get_or_create(
                        user=parent, curso=current_curso,
                        defaults={'rol': CursoMembresia.ROL_APODERADO},
                    )
                messages.success(request, f'Alumno {first_name} {last_name} creado exitosamente.')
            else:
                messages.error(request, 'Nombre y Apellido son requeridos para el alumno.')

        elif action == 'add_gestor':
            puede_gestionar_roles = request.user.is_staff or is_tesorero_principal(request.user, current_curso)
            if not puede_gestionar_roles:
                messages.error(request, 'No tienes permiso para gestionar roles.')
            else:
                gestor_user_id = request.POST.get('gestor_user_id', '').strip()
                rol = request.POST.get('gestor_rol', CursoMembresia.ROL_AYUDANTE)
                if gestor_user_id and current_curso:
                    gestor_user = User.objects.filter(id=gestor_user_id).first()
                    if gestor_user:
                        if rol not in [CursoMembresia.ROL_TESORERO, CursoMembresia.ROL_AYUDANTE]:
                            rol = CursoMembresia.ROL_AYUDANTE
                        # First tesorero of this course becomes principal
                        if rol == CursoMembresia.ROL_TESORERO:
                            ya_tiene_principal = CursoMembresia.objects.filter(
                                curso=current_curso, rol=CursoMembresia.ROL_TESORERO, is_principal=True
                            ).exists()
                        obj, created = CursoMembresia.objects.get_or_create(
                            user=gestor_user, curso=current_curso,
                            defaults={'rol': rol},
                        )
                        if not created:
                            obj.rol = rol
                            obj.save()
                        if rol == CursoMembresia.ROL_TESORERO and not ya_tiene_principal:
                            obj.is_principal = True
                            obj.save()
                        messages.success(request, f'{gestor_user.username} asignado como {obj.get_rol_display()}.')
                    else:
                        messages.error(request, 'Apoderado no encontrado.')

        return redirect('manage_users')

    students = Student.objects.filter(curso=current_curso).select_related('parent').order_by('last_name', 'first_name')
    from django.db.models import Case, When, Value, IntegerField
    membresías = CursoMembresia.objects.filter(
        curso=current_curso,
    ).select_related('user').annotate(
        rol_order=Case(
            When(rol=CursoMembresia.ROL_TESORERO, is_principal=True, then=Value(0)),
            When(rol=CursoMembresia.ROL_AYUDANTE, then=Value(1)),
            When(rol=CursoMembresia.ROL_TESORERO, is_principal=False, then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by('rol_order', 'user__last_name', 'user__first_name')

    es_principal = request.user.is_staff or is_tesorero_principal(request.user, current_curso)

    context = {
        'membresías': membresías,
        'students': students,
        'es_principal': es_principal,
    }
    return render(request, 'fondos/manage_users.html', context)


@gestor_required
def edit_user(request, user_id):
    apoderado = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        new_username = request.POST.get('username', '').strip()
        if new_username and new_username != apoderado.username:
            if User.objects.filter(username=new_username).exists():
                messages.error(request, 'Ese nombre de usuario ya existe.')
                return redirect('manage_users')
            apoderado.username = new_username
        apoderado.first_name = request.POST.get('first_name', apoderado.first_name)
        apoderado.last_name = request.POST.get('last_name', apoderado.last_name)
        apoderado.email = request.POST.get('email', apoderado.email)
        new_password = request.POST.get('password', '').strip()
        if new_password:
            apoderado.set_password(new_password)
        apoderado.save()
        messages.success(request, 'Apoderado actualizado correctamente.')
    return redirect('manage_users')


@gestor_required
def edit_student(request, student_id):
    current_curso = get_current_curso(request)
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        student.first_name = request.POST.get('first_name', student.first_name)
        student.last_name = request.POST.get('last_name', student.last_name)
        parent_id = request.POST.get('parent_id', '')
        new_parent = User.objects.filter(id=parent_id).first() if parent_id else None
        student.parent = new_parent
        student.save()
        if new_parent and current_curso:
            CursoMembresia.objects.get_or_create(
                user=new_parent, curso=current_curso,
                defaults={'rol': CursoMembresia.ROL_APODERADO},
            )
        messages.success(request, 'Alumno actualizado correctamente.')
    return redirect('manage_users')


@gestor_required
def remove_gestor(request, membresia_id):
    if not request.user.is_staff:
        messages.error(request, 'Solo el administrador puede remover gestores.')
        return redirect('manage_users')
    membresia = get_object_or_404(CursoMembresia, id=membresia_id)
    if request.method == 'POST':
        nombre = f"{membresia.user.username} ({membresia.get_rol_display()})"
        membresia.delete()
        messages.success(request, f'{nombre} removido del curso.')
    return redirect('manage_users')


@gestor_required
def promote_to_tesorero(request, user_id):
    current_curso = get_current_curso(request)
    if not (request.user.is_staff or is_tesorero_principal(request.user, current_curso)):
        messages.error(request, 'Solo el administrador o el Tesorero Principal pueden promover a Tesorero.')
        return redirect('manage_users')
    target_user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        membresia, _ = CursoMembresia.objects.get_or_create(
            user=target_user, curso=current_curso,
            defaults={'rol': CursoMembresia.ROL_TESORERO},
        )
        ya_tiene_principal = CursoMembresia.objects.filter(
            curso=current_curso, rol=CursoMembresia.ROL_TESORERO, is_principal=True
        ).exclude(user=target_user).exists()
        membresia.rol = CursoMembresia.ROL_TESORERO
        if not ya_tiene_principal:
            membresia.is_principal = True
        membresia.save()
        messages.success(request, f'{target_user.username} es ahora Tesorero.')
    return redirect('manage_users')


@gestor_required
def revoke_to_apoderado(request, membresia_id):
    current_curso = get_current_curso(request)
    if not (request.user.is_staff or is_tesorero_principal(request.user, current_curso)):
        messages.error(request, 'Solo el administrador o el Tesorero Principal pueden revocar el rol de Tesorero.')
        return redirect('manage_users')
    membresia = get_object_or_404(CursoMembresia, id=membresia_id, curso=current_curso)
    if request.method == 'POST':
        if membresia.is_principal and not request.user.is_staff:
            messages.error(request, 'No puedes revocar al Tesorero Principal.')
            return redirect('manage_users')
        membresia.rol = CursoMembresia.ROL_APODERADO
        membresia.is_principal = False
        membresia.save()
        messages.success(request, f'{membresia.user.username} ha vuelto a ser Apoderado.')
    return redirect('manage_users')


@gestor_required
def apoderado_detail(request, user_id):
    current_curso = get_current_curso(request)
    apoderado = get_object_or_404(User, id=user_id)
    students = Student.objects.filter(parent=apoderado, curso=current_curso).prefetch_related('distributions__activity')
    context = {'apoderado': apoderado, 'students': students}
    return render(request, 'fondos/apoderado_detail.html', context)


@gestor_required
def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    distributions = FundDistribution.objects.filter(student=student).select_related('activity__objetivo').order_by('-activity__date')
    pagos_cuotas = PagoCuota.objects.filter(student=student).select_related('cuota').order_by('-cuota__date')
    meta = student.curso.meta_por_alumno if student.curso else None
    falta = max(0, meta - student.total_funds) if meta else None
    porcentaje = min(100, int(student.total_funds * 100 / meta)) if meta and meta > 0 else None
    context = {
        'student': student,
        'distributions': distributions,
        'pagos_cuotas': pagos_cuotas,
        'meta': meta,
        'falta': falta,
        'porcentaje': porcentaje,
        'objetivos_alumno': _get_objetivos_alumno(student),
    }
    return render(request, 'fondos/student_detail.html', context)


@gestor_required
def cuotas_list(request):
    current_curso = get_current_curso(request)

    if request.method == 'POST':
        name = request.POST.get('name')
        amount = request.POST.get('amount')
        date = request.POST.get('date')
        description = request.POST.get('description', '')
        objetivo_id = request.POST.get('objetivo_id') or None
        if name and amount and date:
            cuota = Cuota.objects.create(
                name=name, amount=int(amount), date=date,
                description=description, curso=current_curso,
                objetivo_id=objetivo_id,
            )
            students = Student.objects.filter(curso=current_curso)
            PagoCuota.objects.bulk_create([
                PagoCuota(student=s, cuota=cuota) for s in students
            ])
            messages.success(request, f'Cuota "{name}" creada para {students.count()} alumnos.')
            return redirect('cuotas_list')

    cuotas = Cuota.objects.filter(curso=current_curso).select_related('objetivo').annotate(
        pagados=Count('pagos', filter=Q(pagos__paid=True)),
        pendientes=Count('pagos', filter=Q(pagos__paid=False)),
    ).order_by('-date')
    objetivos = Objetivo.objects.filter(curso=current_curso)
    context = {'cuotas': cuotas, 'objetivos': objetivos}
    return render(request, 'fondos/cuotas.html', context)


@gestor_required
def delete_cuota(request, cuota_id):
    current_curso = get_current_curso(request)
    cuota = get_object_or_404(Cuota, id=cuota_id, curso=current_curso)
    if request.method == 'POST':
        cuota.delete()
        messages.success(request, f'Cuota "{cuota.name}" eliminada.')
    return redirect('cuotas_list')


@gestor_required
def sync_cuota_students(request, cuota_id):
    current_curso = get_current_curso(request)
    cuota = get_object_or_404(Cuota, id=cuota_id, curso=current_curso)
    if request.method == 'POST':
        existing_ids = cuota.pagos.values_list('student_id', flat=True)
        missing = Student.objects.filter(curso=current_curso).exclude(id__in=existing_ids)
        PagoCuota.objects.bulk_create([PagoCuota(student=s, cuota=cuota) for s in missing])
        if missing:
            messages.success(request, f'{missing.count()} alumno(s) agregado(s) a la cuota.')
        else:
            messages.info(request, 'Todos los alumnos ya estaban en la cuota.')
    return redirect('cuota_detail', cuota_id=cuota_id)


@gestor_required
def edit_cuota(request, cuota_id):
    current_curso = get_current_curso(request)
    cuota = get_object_or_404(Cuota, id=cuota_id, curso=current_curso)
    if request.method == 'POST':
        cuota.name = request.POST.get('name', cuota.name)
        new_date = request.POST.get('date')
        if new_date:
            cuota.date = new_date
        amount_str = request.POST.get('amount', '').strip()
        if amount_str and amount_str.isdigit() and int(amount_str) > 0:
            cuota.amount = int(amount_str)
        cuota.description = request.POST.get('description', cuota.description)
        cuota.objetivo_id = request.POST.get('objetivo_id') or None
        cuota.save()
        messages.success(request, 'Cuota actualizada correctamente.')
    return redirect('cuotas_list')


@gestor_required
def cuota_detail(request, cuota_id):
    current_curso = get_current_curso(request)
    cuota = get_object_or_404(Cuota, id=cuota_id, curso=current_curso)
    filter_status = request.GET.get('filter', 'all')
    sort_by = request.GET.get('sort', 'name')
    sort_order = request.GET.get('order', 'asc')

    pagos = PagoCuota.objects.filter(cuota=cuota).select_related('student__parent').prefetch_related('abonos')

    if filter_status == 'paid':
        pagos = pagos.filter(paid=True)
    elif filter_status == 'pending':
        pagos = pagos.filter(paid=False)

    if sort_by == 'status':
        pagos = pagos.order_by(('paid' if sort_order == 'asc' else '-paid'), 'student__last_name')
    elif sort_by == 'date':
        pagos = pagos.order_by(('paid_date' if sort_order == 'asc' else '-paid_date'))
    else:  # name
        pagos = pagos.order_by(('student__last_name' if sort_order == 'asc' else '-student__last_name'))

    existing_ids = cuota.pagos.values_list('student_id', flat=True)
    alumnos_faltantes = Student.objects.filter(curso=current_curso).exclude(id__in=existing_ids).count()

    context = {
        'cuota': cuota, 'pagos': pagos,
        'alumnos_faltantes': alumnos_faltantes,
        'filter_status': filter_status,
        'sort_by': sort_by, 'sort_order': sort_order,
        'today': timezone.localdate().isoformat(),
    }
    return render(request, 'fondos/cuota_detail.html', context)


@gestor_required
def toggle_pago(request, pago_id):
    current_curso = get_current_curso(request)
    pago = get_object_or_404(PagoCuota, id=pago_id, cuota__curso=current_curso)
    if request.method == 'POST':
        pago.paid = not pago.paid
        if pago.paid:
            from django.utils.dateparse import parse_date
            from datetime import datetime, time
            date_str = request.POST.get('paid_date', '').strip()
            parsed = parse_date(date_str) if date_str else None
            pago.paid_date = timezone.make_aware(datetime.combine(parsed, time())) if parsed else timezone.now()
        else:
            pago.paid_date = None
        pago.save()
    return redirect('cuota_detail', cuota_id=pago.cuota.id)


@gestor_required
def agregar_abono(request, pago_id):
    current_curso = get_current_curso(request)
    pago = get_object_or_404(PagoCuota, id=pago_id, cuota__curso=current_curso)
    if request.method == 'POST':
        amount_str = request.POST.get('amount', '').strip()
        fecha_str = request.POST.get('fecha', '').strip()
        nota = request.POST.get('nota', '').strip()
        from django.utils.dateparse import parse_date
        fecha = parse_date(fecha_str) if fecha_str else timezone.localdate()
        if amount_str and amount_str.isdigit() and int(amount_str) > 0:
            Abono.objects.create(pago=pago, amount=int(amount_str), fecha=fecha, nota=nota)
            # Auto-marcar como pagado si los abonos cubren el total
            if pago.total_abonado >= pago.cuota.amount and not pago.paid:
                pago.paid = True
                pago.paid_date = timezone.now()
                pago.save()
                messages.success(request, f'Abono registrado. Cuota marcada como pagada automáticamente.')
            else:
                messages.success(request, f'Abono de ${int(amount_str):,} registrado.')
        else:
            messages.error(request, 'El monto del abono debe ser un número mayor a 0.')
    return redirect('cuota_detail', cuota_id=pago.cuota.id)


@gestor_required
def editar_abono(request, abono_id):
    current_curso = get_current_curso(request)
    abono = get_object_or_404(Abono, id=abono_id, pago__cuota__curso=current_curso)
    if request.method == 'POST':
        amount_str = request.POST.get('amount', '').strip()
        fecha_str = request.POST.get('fecha', '').strip()
        nota = request.POST.get('nota', '').strip()
        from django.utils.dateparse import parse_date
        if amount_str and amount_str.isdigit() and int(amount_str) > 0:
            abono.amount = int(amount_str)
            if fecha_str:
                abono.fecha = parse_date(fecha_str) or abono.fecha
            abono.nota = nota
            abono.save()
            # Recalcular estado del pago
            pago = abono.pago
            if pago.total_abonado >= pago.cuota.amount and not pago.paid:
                pago.paid = True
                pago.paid_date = timezone.now()
                pago.save()
            elif pago.total_abonado < pago.cuota.amount and pago.paid:
                pago.paid = False
                pago.paid_date = None
                pago.save()
            messages.success(request, 'Abono actualizado.')
        else:
            messages.error(request, 'Monto inválido.')
    return redirect('cuota_detail', cuota_id=abono.pago.cuota.id)


@gestor_required
def eliminar_abono(request, abono_id):
    current_curso = get_current_curso(request)
    abono = get_object_or_404(Abono, id=abono_id, pago__cuota__curso=current_curso)
    cuota_id = abono.pago.cuota.id
    pago = abono.pago
    if request.method == 'POST':
        abono.delete()
        # Si ya no cubre el total, revertir a pendiente
        if pago.total_abonado < pago.cuota.amount and pago.paid:
            pago.paid = False
            pago.paid_date = None
            pago.save()
        messages.success(request, 'Abono eliminado.')
    return redirect('cuota_detail', cuota_id=cuota_id)


@gestor_required
def delete_activity(request, activity_id):
    current_curso = get_current_curso(request)
    activity = get_object_or_404(Activity, id=activity_id, curso=current_curso)
    if request.method == 'POST':
        name = activity.name
        activity.delete()
        messages.success(request, f'Actividad "{name}" eliminada.')
    return redirect('admin_dashboard')


@gestor_required
def clear_distributions(request, activity_id):
    current_curso = get_current_curso(request)
    activity = get_object_or_404(Activity, id=activity_id, curso=current_curso)
    if request.method == 'POST':
        count = activity.distributions.count()
        activity.distributions.all().delete()
        messages.success(request, f'{count} distribución(es) eliminadas. Puedes volver a repartir.')
    return redirect('distribute_funds', activity_id=activity_id)


@gestor_required
def edit_activity(request, activity_id):
    current_curso = get_current_curso(request)
    activity = get_object_or_404(Activity, id=activity_id, curso=current_curso)
    if request.method == 'POST':
        activity.name = request.POST.get('name', activity.name)
        new_date = request.POST.get('date')
        if new_date:
            activity.date = new_date
        activity.total_amount = request.POST.get('total_amount', activity.total_amount)
        activity.description = request.POST.get('description', activity.description)
        objetivo_id = request.POST.get('objetivo_id') or None
        activity.objetivo_id = objetivo_id
        activity.save()
        messages.success(request, 'Actividad actualizada correctamente.')
    return redirect('admin_dashboard')


@gestor_required
def edit_distribution(request, dist_id):
    current_curso = get_current_curso(request)
    dist = get_object_or_404(FundDistribution, id=dist_id, activity__curso=current_curso)
    if request.method == 'POST':
        new_amount = request.POST.get('amount')
        if new_amount and new_amount.lstrip('-').isdigit():
            dist.amount = int(new_amount)
            dist.save()
            messages.success(request, 'Monto actualizado correctamente.')
        else:
            messages.error(request, 'Monto numérico inválido.')
    return redirect('distribute_funds', activity_id=dist.activity.id)


@gestor_required
def delete_distribution(request, dist_id):
    current_curso = get_current_curso(request)
    dist = get_object_or_404(FundDistribution, id=dist_id, activity__curso=current_curso)
    activity_id = dist.activity.id
    if request.method == 'POST':
        dist.delete()
        messages.success(request, 'Distribución de fondo eliminada.')
    return redirect('distribute_funds', activity_id=activity_id)


@gestor_required
def reportes(request):
    current_curso = get_current_curso(request)
    return render(request, 'fondos/reportes.html', {'curso': current_curso})


@gestor_required
def generar_reporte(request):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    current_curso = get_current_curso(request)
    if not current_curso.es_avanzado_activo:
        messages.error(request, 'Esta función requiere un curso con plan Avanzado activo. Verifica el período configurado.')
        return redirect('reportes')

    tipo_reporte = request.POST.get('tipo_reporte', 'actividades')
    formato = request.POST.get('formato', 'excel')

    if formato == 'pdf':
        return _generar_reporte_pdf(request, current_curso, tipo_reporte)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    color_header = 'FF4F46E5'
    color_subheader = 'FFE8E7FF'
    color_pendiente = 'FFFFF3CD'
    color_pagado = 'FFD4EDDA'

    def style_header(cell, bg=color_header):
        cell.font = Font(bold=True, color='FFFFFFFF' if bg == color_header else 'FF000000')
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def style_title(cell):
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    thin = Side(style='thin', color='FFD0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── REPORTE: ACTIVIDADES ──────────────────────────────────────────────────
    if tipo_reporte in ('actividades', 'ambos'):
        ws = wb.create_sheet('Actividades')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

        row = 1
        activities = Activity.objects.filter(curso=current_curso).order_by('-date').prefetch_related('distributions__student__parent')

        for act in activities:
            # Activity title
            ws.cell(row, 1, f'{act.name}  —  {act.date.strftime("%d/%m/%Y")}  —  Ref: ${act.total_amount}')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            style_title(ws.cell(row, 1))
            row += 1

            # Column headers
            for col, label in enumerate(['Alumno', 'Apoderado', 'Monto Asignado'], 1):
                c = ws.cell(row, col, label)
                style_header(c)
                c.border = border
            row += 1

            dists = act.distributions.select_related('student__parent').order_by('student__last_name')
            total = 0
            for dist in dists:
                ws.cell(row, 1, f'{dist.student.last_name} {dist.student.first_name}').border = border
                parent = dist.student.parent
                ws.cell(row, 2, f'{parent.last_name} {parent.first_name}' if parent else '—').border = border
                ws.cell(row, 3, dist.amount).border = border
                ws.cell(row, 3).number_format = '"$"#,##0'
                total += dist.amount
                row += 1

            if dists:
                ws.cell(row, 2, 'Total repartido').font = Font(bold=True)
                ws.cell(row, 3, total).font = Font(bold=True)
                ws.cell(row, 3).number_format = '"$"#,##0'
            row += 2  # blank separator

    # ── REPORTE: CUOTAS ───────────────────────────────────────────────────────
    if tipo_reporte in ('cuotas', 'ambos'):
        ws = wb.create_sheet('Cuotas')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 18

        row = 1
        cuotas = Cuota.objects.filter(curso=current_curso).order_by('-date').prefetch_related('pagos__student__parent')

        for cuota in cuotas:
            ws.cell(row, 1, f'{cuota.name}  —  {cuota.date.strftime("%d/%m/%Y")}  —  ${cuota.amount} por alumno')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            style_title(ws.cell(row, 1))
            row += 1

            for col, label in enumerate(['Alumno', 'Apoderado', 'Estado', 'Fecha de Pago'], 1):
                c = ws.cell(row, col, label)
                style_header(c)
                c.border = border
            row += 1

            pagos = cuota.pagos.select_related('student__parent').order_by('paid', 'student__last_name')
            pagados = pendientes = 0
            for pago in pagos:
                fill_color = color_pagado if pago.paid else color_pendiente
                fill = PatternFill('solid', fgColor=fill_color)
                estado = 'Pagado' if pago.paid else 'Pendiente'
                fecha = pago.paid_date.strftime('%d/%m/%Y') if pago.paid_date else '—'

                for col, val in enumerate([
                    f'{pago.student.last_name} {pago.student.first_name}',
                    f'{pago.student.parent.last_name} {pago.student.parent.first_name}' if pago.student.parent else '—',
                    estado, fecha,
                ], 1):
                    c = ws.cell(row, col, val)
                    c.fill = fill
                    c.border = border
                if pago.paid:
                    pagados += 1
                else:
                    pendientes += 1
                row += 1

            ws.cell(row, 2, 'Resumen').font = Font(bold=True)
            ws.cell(row, 3, f'{pagados} pagados / {pendientes} pendientes').font = Font(bold=True)
            row += 2

    # ── REPORTE: RESUMEN GENERAL ───────────────────────────────────────────────
    if tipo_reporte == 'resumen':
        ws = wb.create_sheet('Resumen General')
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16

        ws.cell(1, 1, f'Resumen General — {current_curso}')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        style_title(ws.cell(1, 1))

        for col, label in enumerate(['Alumno', 'Apoderado', 'Total Actividades', 'Total Cuotas Pagadas', 'Total Fondos'], 1):
            c = ws.cell(2, col, label)
            style_header(c)
            c.border = border

        students = Student.objects.filter(curso=current_curso).select_related('parent').order_by('last_name', 'first_name')
        for i, student in enumerate(students, 3):
            ws.cell(i, 1, f'{student.last_name} {student.first_name}').border = border
            parent = student.parent
            ws.cell(i, 2, f'{parent.last_name} {parent.first_name}' if parent else '—').border = border
            ws.cell(i, 3, student.total_actividades).border = border
            ws.cell(i, 3).number_format = '"$"#,##0'
            ws.cell(i, 4, student.total_cuotas_pagadas).border = border
            ws.cell(i, 4).number_format = '"$"#,##0'
            ws.cell(i, 5, student.total_funds).border = border
            ws.cell(i, 5).number_format = '"$"#,##0'

    if not wb.sheetnames:
        wb.create_sheet('Sin datos')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_archivo = f'reporte_{current_curso.name.replace(" ", "_")}_{current_curso.year}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


def _generar_reporte_pdf(request, current_curso, tipo_reporte):
    import io
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

    COLOR_HEADER = colors.HexColor('#4F46E5')
    COLOR_SUBHEADER = colors.HexColor('#E8E7FF')
    COLOR_PAGADO = colors.HexColor('#D4EDDA')
    COLOR_PENDIENTE = colors.HexColor('#FFF3CD')
    COLOR_TOTAL = colors.HexColor('#F1F3F5')

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('titulo', parent=styles['Heading2'], textColor=COLOR_HEADER, spaceAfter=4)
    subtitulo_style = ParagraphStyle('subtitulo', parent=styles['Heading3'], spaceAfter=2)
    normal = styles['Normal']

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f'Reporte {current_curso}',
    )

    story = []

    def header_row_style(n_cols):
        return [
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]

    # Título general
    story.append(Paragraph(f'Reporte — {current_curso}', titulo_style))
    story.append(Spacer(1, 0.3 * cm))

    # ── ACTIVIDADES ───────────────────────────────────────────────────────────
    if tipo_reporte in ('actividades', 'ambos'):
        story.append(Paragraph('Actividades', subtitulo_style))
        activities = Activity.objects.filter(curso=current_curso).order_by('-date').prefetch_related('distributions__student__parent')

        for act in activities:
            story.append(Paragraph(
                f'<b>{act.name}</b>  ·  {act.date.strftime("%d/%m/%Y")}  ·  Ref: ${act.total_amount:,}',
                normal,
            ))
            dists = act.distributions.select_related('student__parent').order_by('student__last_name')
            data = [['Alumno', 'Apoderado', 'Monto Asignado']]
            total = 0
            for dist in dists:
                parent = dist.student.parent
                data.append([
                    f'{dist.student.last_name} {dist.student.first_name}',
                    f'{parent.last_name} {parent.first_name}' if parent else '—',
                    f'${dist.amount:,}',
                ])
                total += dist.amount
            if dists:
                data.append(['', 'Total repartido', f'${total:,}'])

            t = Table(data, colWidths=[8 * cm, 8 * cm, 4 * cm])
            ts = header_row_style(3)
            if dists:
                ts += [
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, -1), (-1, -1), COLOR_TOTAL),
                ]
            t.setStyle(TableStyle(ts))
            story.append(Spacer(1, 0.2 * cm))
            story.append(t)
            story.append(Spacer(1, 0.5 * cm))

        if tipo_reporte == 'ambos':
            story.append(PageBreak())

    # ── CUOTAS ────────────────────────────────────────────────────────────────
    if tipo_reporte in ('cuotas', 'ambos'):
        story.append(Paragraph('Cuotas', subtitulo_style))
        cuotas = Cuota.objects.filter(curso=current_curso).order_by('-date').prefetch_related('pagos__student__parent')

        for cuota in cuotas:
            story.append(Paragraph(
                f'<b>{cuota.name}</b>  ·  {cuota.date.strftime("%d/%m/%Y")}  ·  ${cuota.amount:,} por alumno',
                normal,
            ))
            pagos = cuota.pagos.select_related('student__parent').order_by('paid', 'student__last_name')
            data = [['Alumno', 'Apoderado', 'Estado', 'Fecha de Pago']]
            pagados = pendientes = 0
            row_colors = []
            for i, pago in enumerate(pagos, 1):
                estado = 'Pagado' if pago.paid else 'Pendiente'
                fecha = pago.paid_date.strftime('%d/%m/%Y') if pago.paid_date else '—'
                parent = pago.student.parent
                data.append([
                    f'{pago.student.last_name} {pago.student.first_name}',
                    f'{parent.last_name} {parent.first_name}' if parent else '—',
                    estado, fecha,
                ])
                row_colors.append(COLOR_PAGADO if pago.paid else COLOR_PENDIENTE)
                if pago.paid:
                    pagados += 1
                else:
                    pendientes += 1
            data.append(['', f'{pagados} pagados / {pendientes} pendientes', '', ''])

            t = Table(data, colWidths=[7 * cm, 7 * cm, 3 * cm, 4 * cm])
            ts = header_row_style(4)
            # override row backgrounds with paid/pending colors
            ts = [s for s in ts if s[0] != 'ROWBACKGROUNDS']
            for i, bg in enumerate(row_colors, 1):
                ts.append(('BACKGROUND', (0, i), (-1, i), bg))
            ts += [
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), COLOR_TOTAL),
            ]
            t.setStyle(TableStyle(ts))
            story.append(Spacer(1, 0.2 * cm))
            story.append(t)
            story.append(Spacer(1, 0.5 * cm))

    # ── RESUMEN GENERAL ───────────────────────────────────────────────────────
    if tipo_reporte == 'resumen':
        story.append(Paragraph('Resumen General', subtitulo_style))
        students = Student.objects.filter(curso=current_curso).select_related('parent').order_by('last_name', 'first_name')
        data = [['Alumno', 'Apoderado', 'Total Actividades', 'Total Cuotas Pagadas', 'Total Fondos']]
        for student in students:
            parent = student.parent
            data.append([
                f'{student.last_name} {student.first_name}',
                f'{parent.last_name} {parent.first_name}' if parent else '—',
                f'${student.total_actividades:,}',
                f'${student.total_cuotas_pagadas:,}',
                f'${student.total_funds:,}',
            ])
        t = Table(data, colWidths=[7 * cm, 7 * cm, 4 * cm, 4 * cm, 4 * cm])
        t.setStyle(TableStyle(header_row_style(5)))
        story.append(t)

    doc.build(story)
    buf.seek(0)
    nombre_archivo = f'reporte_{current_curso.name.replace(" ", "_")}_{current_curso.year}.pdf'
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ─── OBJETIVOS ────────────────────────────────────────────────────────────────

def _get_objetivos_alumno(student):
    """Devuelve lista de dicts con el aporte y meta individual del alumno por cada objetivo."""
    if not student.curso:
        return []
    objetivos = Objetivo.objects.filter(curso=student.curso)
    result = []
    for obj in objetivos:
        aporte_act = FundDistribution.objects.filter(
            student=student, activity__objetivo=obj
        ).aggregate(total=Sum('amount'))['total'] or 0
        aporte_cuotas = PagoCuota.objects.filter(
            student=student, cuota__objetivo=obj, paid=True
        ).aggregate(total=Sum('cuota__amount'))['total'] or 0
        total = aporte_act + aporte_cuotas
        # Meta individual del alumno para este objetivo
        try:
            entrada = ObjetivoAlumno.objects.get(objetivo=obj, student=student)
            meta = entrada.meta_alumno
        except ObjetivoAlumno.DoesNotExist:
            meta = obj.monto_meta
        porcentaje = min(100, int(total * 100 / meta)) if meta and meta > 0 else None
        falta = max(0, meta - total) if meta else None
        result.append({
            'obj': obj,
            'total': total,
            'aporte_act': aporte_act,
            'aporte_cuotas': aporte_cuotas,
            'meta': meta,
            'porcentaje': porcentaje,
            'falta': falta,
        })
    return result


@gestor_required
def objetivos(request):
    current_curso = get_current_curso(request)

    students_curso = Student.objects.filter(curso=current_curso).order_by('last_name', 'first_name')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            name = request.POST.get('name', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            meta_str = request.POST.get('monto_meta', '').strip()
            if name:
                obj = Objetivo.objects.create(
                    curso=current_curso, name=name, descripcion=descripcion,
                    monto_meta=int(meta_str) if meta_str.isdigit() else None,
                )
                # Auto-crear entradas para todos los alumnos con multiplicador=1
                ObjetivoAlumno.objects.bulk_create([
                    ObjetivoAlumno(objetivo=obj, student=s, multiplicador=1)
                    for s in students_curso
                ])
                messages.success(request, f'Objetivo "{name}" creado para {students_curso.count()} alumno(s).')
            else:
                messages.error(request, 'El nombre es requerido.')
        elif action == 'edit':
            obj_id = request.POST.get('objetivo_id')
            obj = get_object_or_404(Objetivo, id=obj_id, curso=current_curso)
            obj.name = request.POST.get('name', obj.name).strip()
            obj.descripcion = request.POST.get('descripcion', '').strip()
            meta_str = request.POST.get('monto_meta', '').strip()
            obj.monto_meta = int(meta_str) if meta_str.isdigit() else None
            obj.save()
            messages.success(request, 'Objetivo actualizado.')
        elif action == 'set_multiplicador':
            # Actualizar multiplicadores por alumno para un objetivo
            obj_id = request.POST.get('objetivo_id')
            obj = get_object_or_404(Objetivo, id=obj_id, curso=current_curso)
            for s in students_curso:
                mult_str = request.POST.get(f'mult_{s.id}', '1').strip()
                mult = int(mult_str) if mult_str.isdigit() and int(mult_str) >= 1 else 1
                ObjetivoAlumno.objects.update_or_create(
                    objetivo=obj, student=s,
                    defaults={'multiplicador': mult},
                )
            messages.success(request, 'Multiplicadores actualizados.')
        elif action == 'sync_alumnos':
            # Agregar alumnos nuevos que no tengan entrada aún
            obj_id = request.POST.get('objetivo_id')
            obj = get_object_or_404(Objetivo, id=obj_id, curso=current_curso)
            existing_ids = obj.alumnos_meta.values_list('student_id', flat=True)
            nuevos = students_curso.exclude(id__in=existing_ids)
            ObjetivoAlumno.objects.bulk_create([
                ObjetivoAlumno(objetivo=obj, student=s, multiplicador=1) for s in nuevos
            ])
            if nuevos:
                messages.success(request, f'{nuevos.count()} alumno(s) agregado(s) al objetivo.')
            else:
                messages.info(request, 'Todos los alumnos ya están registrados.')
        elif action == 'delete':
            obj_id = request.POST.get('objetivo_id')
            obj = get_object_or_404(Objetivo, id=obj_id, curso=current_curso)
            obj.delete()
            messages.success(request, 'Objetivo eliminado.')
        return redirect('objetivos')

    # Calcular ingresos y gastos por objetivo
    all_objetivos = Objetivo.objects.filter(curso=current_curso).prefetch_related(
        'alumnos_meta__student', 'activities', 'cuotas', 'gastos'
    )
    objetivos_data = []
    for obj in all_objetivos:
        ingresos_act = sum(a.total_amount for a in obj.activities.all())
        ingresos_cuotas = sum(c.total_recaudado for c in obj.cuotas.all())
        total_ingresos = ingresos_act + ingresos_cuotas
        total_gastos = sum(g.amount for g in obj.gastos.all())
        balance = total_ingresos - total_gastos
        meta_total = obj.monto_meta_total
        porcentaje = min(100, int(balance * 100 / meta_total)) if meta_total and meta_total > 0 else None
        # Alumnos faltantes (no tienen entrada)
        existing_ids = [e.student_id for e in obj.alumnos_meta.all()]
        faltantes = sum(1 for s in students_curso if s.id not in existing_ids)
        objetivos_data.append({
            'obj': obj,
            'ingresos_act': ingresos_act,
            'ingresos_cuotas': ingresos_cuotas,
            'total_ingresos': total_ingresos,
            'total_gastos': total_gastos,
            'balance': balance,
            'meta_total': meta_total,
            'porcentaje': porcentaje,
            'faltantes': faltantes,
        })

    # Totales del curso (sin filtrar por objetivo)
    total_ingresos_curso = (
        sum(a.total_amount for a in Activity.objects.filter(curso=current_curso)) +
        sum(c.total_recaudado for c in Cuota.objects.filter(curso=current_curso))
    )
    total_gastos_curso = sum(g.amount for g in Gasto.objects.filter(curso=current_curso))

    context = {
        'objetivos_data': objetivos_data,
        'total_ingresos_curso': total_ingresos_curso,
        'total_gastos_curso': total_gastos_curso,
        'balance_curso': total_ingresos_curso - total_gastos_curso,
    }
    return render(request, 'fondos/objetivos.html', context)


# ─── GASTOS ───────────────────────────────────────────────────────────────────

@gestor_required
def gastos_list(request):
    current_curso = get_current_curso(request)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            name = request.POST.get('name', '').strip()
            amount_str = request.POST.get('amount', '').strip()
            date_str = request.POST.get('date', '').strip()
            nota = request.POST.get('nota', '').strip()
            objetivo_id = request.POST.get('objetivo_id') or None
            if name and amount_str.isdigit() and int(amount_str) > 0 and date_str:
                Gasto.objects.create(
                    curso=current_curso, name=name, amount=int(amount_str),
                    date=date_str, nota=nota, objetivo_id=objetivo_id,
                )
                messages.success(request, f'Gasto "{name}" registrado.')
            else:
                messages.error(request, 'Nombre, monto y fecha son requeridos.')
        elif action == 'edit':
            gasto_id = request.POST.get('gasto_id')
            gasto = get_object_or_404(Gasto, id=gasto_id, curso=current_curso)
            gasto.name = request.POST.get('name', gasto.name).strip()
            amount_str = request.POST.get('amount', '').strip()
            if amount_str.isdigit() and int(amount_str) > 0:
                gasto.amount = int(amount_str)
            date_str = request.POST.get('date', '').strip()
            if date_str:
                gasto.date = date_str
            gasto.nota = request.POST.get('nota', '').strip()
            gasto.objetivo_id = request.POST.get('objetivo_id') or None
            gasto.save()
            messages.success(request, 'Gasto actualizado.')
        elif action == 'delete':
            gasto_id = request.POST.get('gasto_id')
            gasto = get_object_or_404(Gasto, id=gasto_id, curso=current_curso)
            gasto.delete()
            messages.success(request, 'Gasto eliminado.')
        return redirect('gastos_list')

    gastos = Gasto.objects.filter(curso=current_curso).select_related('objetivo').order_by('-date')
    objetivos = Objetivo.objects.filter(curso=current_curso)
    total_gastos = sum(g.amount for g in gastos)
    context = {'gastos': gastos, 'objetivos': objetivos, 'total_gastos': total_gastos,
               'today': timezone.localdate().isoformat()}
    return render(request, 'fondos/gastos.html', context)


@gestor_required
def reporte_apoderados(request):
    current_curso = get_current_curso(request)

    apoderado_ids = CursoMembresia.objects.filter(
        curso=current_curso, rol=CursoMembresia.ROL_APODERADO,
    ).values_list('user_id', flat=True)

    apoderados = User.objects.filter(id__in=apoderado_ids).order_by('last_name', 'first_name')

    apoderados_data = []
    for apoderado in apoderados:
        students = Student.objects.filter(
            parent=apoderado, curso=current_curso,
        ).prefetch_related('distributions__activity', 'pagos__cuota')
        apoderados_data.append({
            'apoderado': apoderado,
            'students': students,
        })

    sin_apoderado = Student.objects.filter(curso=current_curso, parent__isnull=True)

    context = {
        'apoderados_data': apoderados_data,
        'sin_apoderado': sin_apoderado,
    }
    return render(request, 'fondos/reporte_apoderados.html', context)


@login_required
def admin_global_dashboard(request):
    """Dashboard global solo para administradores (is_staff).
    Muestra métricas globales y el estado del plan avanzado de cada curso."""
    if not request.user.is_staff:
        messages.error(request, 'Acceso denegado.')
        return redirect('dashboard')

    total_cursos = Curso.objects.count()
    total_alumnos = Student.objects.count()
    # Apoderados: todos los usuarios con membresía en algún curso (cualquier rol)
    total_apoderados = User.objects.filter(
        membresias__isnull=False
    ).distinct().count()

    cursos = Curso.objects.annotate(
        num_alumnos=Count('students', distinct=True),
        # Todos los miembros del curso (apoderados + tesoreros + ayudantes)
        num_apoderados=Count('membresias', distinct=True),
    ).order_by('-year', 'name')

    context = {
        'total_cursos': total_cursos,
        'total_alumnos': total_alumnos,
        'total_apoderados': total_apoderados,
        'cursos': cursos,
        'today': timezone.localdate(),
    }
    return render(request, 'fondos/admin_global_dashboard.html', context)


@login_required
def set_avanzado_periodo(request, curso_id):
    """Permite al administrador (is_staff) configurar el período de plan avanzado de un curso."""
    if not request.user.is_staff:
        messages.error(request, 'Acceso denegado.')
        return redirect('dashboard')

    curso = get_object_or_404(Curso, id=curso_id)

    if request.method == 'POST':
        from django.utils.dateparse import parse_date
        desde_str = request.POST.get('avanzado_desde', '').strip()
        hasta_str = request.POST.get('avanzado_hasta', '').strip()

        desde = parse_date(desde_str) if desde_str else None
        hasta = parse_date(hasta_str) if hasta_str else None

        if desde and hasta and hasta < desde:
            messages.error(request, 'La fecha de fin no puede ser anterior a la fecha de inicio.')
        else:
            curso.avanzado_desde = desde
            curso.avanzado_hasta = hasta
            curso.save()
            if desde and hasta:
                messages.success(
                    request,
                    f'Período avanzado configurado para {curso}: {desde.strftime("%d/%m/%Y")} — {hasta.strftime("%d/%m/%Y")}.'
                )
            else:
                messages.success(request, f'Período avanzado eliminado para {curso}.')

    return redirect('admin_global_dashboard')
