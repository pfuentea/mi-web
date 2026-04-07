"""
Convierte un archivo Excel (.xlsx / .xls) a CSV.

Uso:
    python scripts/excel_to_csv.py
    python scripts/excel_to_csv.py archivo.xlsx
    python scripts/excel_to_csv.py archivo.xlsx -o salida.csv
    python scripts/excel_to_csv.py archivo.xlsx -s "Hoja2"
    python scripts/excel_to_csv.py archivo.xlsx --todas-las-hojas

Requiere:
    pip install openpyxl
"""

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: falta instalar openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)


def excel_a_csv(excel_path: Path, csv_path: Path, nombre_hoja: str | None = None):
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if nombre_hoja:
        if nombre_hoja not in wb.sheetnames:
            print(f"Error: la hoja '{nombre_hoja}' no existe. Hojas disponibles: {wb.sheetnames}")
            sys.exit(1)
        hoja = wb[nombre_hoja]
    else:
        hoja = wb.active

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for fila in hoja.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in fila])

    print(f"OK '{hoja.title}' -> {csv_path}  ({hoja.max_row} filas, {hoja.max_column} columnas)")


def main():
    # Ruta base del proyecto (un nivel arriba de /scripts)
    base_dir = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(description="Convierte Excel a CSV")
    parser.add_argument(
        "excel",
        nargs="?",
        default=str(base_dir / "Listado.xlsx"),
        help="Ruta al archivo Excel (default: Listado.xlsx)",
    )
    parser.add_argument("-o", "--output", help="Ruta del CSV de salida (default: mismo nombre que el Excel)")
    parser.add_argument("-s", "--hoja", help="Nombre de la hoja a exportar (default: hoja activa)")
    parser.add_argument(
        "--todas-las-hojas",
        action="store_true",
        help="Exportar todas las hojas como CSVs separados",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"Error: no se encontró el archivo '{excel_path}'")
        sys.exit(1)

    if args.todas_las_hojas:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        for nombre in wb.sheetnames:
            csv_path = excel_path.with_name(f"{excel_path.stem}_{nombre}.csv")
            excel_a_csv(excel_path, csv_path, nombre_hoja=nombre)
    else:
        if args.output:
            csv_path = Path(args.output)
        else:
            csv_path = excel_path.with_suffix(".csv")
        excel_a_csv(excel_path, csv_path, nombre_hoja=args.hoja)


if __name__ == "__main__":
    main()
